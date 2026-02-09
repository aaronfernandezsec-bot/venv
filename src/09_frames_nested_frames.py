# 09_frames_nested_frames.py
from pathlib import Path
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options

from openpyxl import Workbook

BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUTS_DIR = BASE_DIR / "outputs"
OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)

def run_tag():
    # mismo estilo que tus evidencias: fecha_hora + consecutivo
    return datetime.now().strftime("%Y%m%d_%H%M%S") + "_009"

opts = Options()
opts.add_argument("--start-maximized")

driver = webdriver.Chrome(options=opts)
wait = WebDriverWait(driver, 15)

try:
    driver.get("https://the-internet.herokuapp.com/nested_frames")

    # 1) Ir al frame BOTTOM y obtener texto
    wait.until(EC.presence_of_element_located((By.TAG_NAME, "frameset")))
    driver.switch_to.frame("frame-bottom")

    body = wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
    bottom_text = body.text.strip()

    # 2) Highlight visual del texto (en el frame BOTTOM)
    driver.execute_script(
        """
        const el = arguments[0];
        el.style.background = '#fff3a0';
        el.style.border = '3px solid #ff3b30';
        el.style.padding = '10px';
        el.style.borderRadius = '8px';
        """,
        body
    )

    tag = run_tag()

    # 3) Screenshot con highlight
    img_path = OUTPUTS_DIR / f"{tag}_frames_bottom_highlight.png"
    driver.save_screenshot(str(img_path))

    # 4) Exportar a Excel el texto BOTTOM
    xlsx_path = OUTPUTS_DIR / f"{tag}_frames_bottom.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "BOTTOM"
    ws["A1"] = "Texto (BOTTOM)"
    ws["A2"] = bottom_text
    wb.save(str(xlsx_path))

    print("OK: BOTTOM extraído y exportado")
    print("TEXTO:", bottom_text)
    print("EVIDENCIAS:")
    print(" -", img_path)
    print(" -", xlsx_path)

finally:
    driver.quit()

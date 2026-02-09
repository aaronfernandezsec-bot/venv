from pathlib import Path
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import Workbook


# ========= Config base (mismo estilo que traes) =========
BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUTS_DIR = BASE_DIR / "outputs"
OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)

RUN_TS = datetime.now().strftime("%Y%m%d_%H%M%S")
OUT_XLSX = OUTPUTS_DIR / f"{RUN_TS}_010_large_deep_dom_12_1.xlsx"
OUT_PNG = OUTPUTS_DIR / f"{RUN_TS}_010_large_deep_dom_evidence.png"

URL = "https://the-internet.herokuapp.com/large"


def build_driver() -> webdriver.Chrome:
    opts = webdriver.ChromeOptions()
    opts.add_argument("--start-maximized")
    # Nota: NO mobile emulation (no lo necesitamos aquí)
    return webdriver.Chrome(options=opts)


driver = build_driver()
wait = WebDriverWait(driver, 20)

try:
    driver.get(URL)

    # Espera a que exista al menos un "12.1" en la página
    wait.until(EC.presence_of_element_located((By.XPATH, "//*[normalize-space(.)='12.1']")))

    # 1) 12.1 de la LISTA (normalmente es <li>)
    list_121 = wait.until(
        EC.presence_of_element_located((By.XPATH, "//ul//li[normalize-space(.)='12.1']"))
    ).text.strip()

    # 2) 12.1 de la TABLA (normalmente es <td>)
    table_121 = wait.until(
        EC.presence_of_element_located((By.XPATH, "//table//td[normalize-space(.)='12.1']"))
    ).text.strip()

    # Evidencia visual (screenshot)
    driver.save_screenshot(str(OUT_PNG))

    # Export a Excel con ambos valores
    wb = Workbook()
    ws = wb.active
    ws.title = "Large&DeepDOM"

    ws["A1"] = "Fuente"
    ws["B1"] = "Valor"

    ws["A2"] = "Lista"
    ws["B2"] = list_121

    ws["A3"] = "Tabla"
    ws["B3"] = table_121

    wb.save(str(OUT_XLSX))

    print("OK: extraído 12.1 (lista y tabla) y exportado a Excel")
    print("EVIDENCIA PNG:", OUT_PNG)
    print("EVIDENCIA XLSX:", OUT_XLSX)

finally:
    driver.quit()
